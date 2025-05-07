import boto3
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================
# Forecast JSON (Fixed structure)
# ==========================
FORECAST_JSON = {
    "region": "Cirencester",
    "forecast": [
        {
            "year": 2025,
            "predicted_demand": 5200,
            "predicted_supply": 3900,
            "demand_supply_gap": 1300,
            "population": 552000,
            "housing_stock": 243000,
            "net_migration": 2200,
            "bungalow_demand": 850,
            "bungalow_supply": 400,
            "population_makeup": {
                "age_distribution": {
                    "0-14": 82000,
                    "15-24": 62000,
                    "25-44": 138000,
                    "45-64": 142000,
                    "65+": 128000
                },
                "household_types": {
                    "single_person": 72000,
                    "couple_no_children": 58000,
                    "couple_with_children": 87000,
                    "single_parent": 39000,
                    "other": 24000
                },
                "tenure": {
                    "owner_occupied": 64,
                    "private_rented": 20,
                    "social_rented": 16
                },
                "ethnicity": {
                    "White": 91,
                    "Asian": 5,
                    "Black": 2,
                    "Mixed": 1,
                    "Other": 1
                }
            }
        },
        {
            "year": 2026,
            "predicted_demand": 5400,
            "predicted_supply": 4050,
            "demand_supply_gap": 1350,
            "population": 558500,
            "housing_stock": 247200,
            "net_migration": 2300,
            "bungalow_demand": 900,
            "bungalow_supply": 420,
            "population_makeup": {
                "age_distribution": {
                    "0-14": 82500,
                    "15-24": 63000,
                    "25-44": 139500,
                    "45-64": 143000,
                    "65+": 130500
                },
                "household_types": {
                    "single_person": 73000,
                    "couple_no_children": 58500,
                    "couple_with_children": 88000,
                    "single_parent": 39500,
                    "other": 24500
                },
                "tenure": {
                    "owner_occupied": 63.5,
                    "private_rented": 20.5,
                    "social_rented": 16
                },
                "ethnicity": {
                    "White": 90.8,
                    "Asian": 5.1,
                    "Black": 2.1,
                    "Mixed": 1,
                    "Other": 1
                }
            }
        },
        {
            "year": 2027,
            "predicted_demand": 5600,
            "predicted_supply": 4200,
            "demand_supply_gap": 1400,
            "population": 565200,
            "housing_stock": 251500,
            "net_migration": 2400,
            "bungalow_demand": 950,
            "bungalow_supply": 450,
            "population_makeup": {
                "age_distribution": {
                    "0-14": 83000,
                    "15-24": 64000,
                    "25-44": 141000,
                    "45-64": 144000,
                    "65+": 133200
                },
                "household_types": {
                    "single_person": 74000,
                    "couple_no_children": 59000,
                    "couple_with_children": 89000,
                    "single_parent": 40000,
                    "other": 25000
                },
                "tenure": {
                    "owner_occupied": 63,
                    "private_rented": 21,
                    "social_rented": 16
                },
                "ethnicity": {
                    "White": 90.6,
                    "Asian": 5.2,
                    "Black": 2.2,
                    "Mixed": 1,
                    "Other": 1
                }
            }
        },
        {
            "year": 2028,
            "predicted_demand": 5800,
            "predicted_supply": 4350,
            "demand_supply_gap": 1450,
            "population": 572000,
            "housing_stock": 255900,
            "net_migration": 2500,
            "bungalow_demand": 1000,
            "bungalow_supply": 470,
            "population_makeup": {
                "age_distribution": {
                    "0-14": 83500,
                    "15-24": 65000,
                    "25-44": 142000,
                    "45-64": 145000,
                    "65+": 135500
                },
                "household_types": {
                    "single_person": 75000,
                    "couple_no_children": 59500,
                    "couple_with_children": 90000,
                    "single_parent": 40500,
                    "other": 25500
                },
                "tenure": {
                    "owner_occupied": 62.5,
                    "private_rented": 21.5,
                    "social_rented": 16
                },
                "ethnicity": {
                    "White": 90.4,
                    "Asian": 5.3,
                    "Black": 2.3,
                    "Mixed": 1,
                    "Other": 1
                }
            }
        },
        {
            "year": 2029,
            "predicted_demand": 6000,
            "predicted_supply": 4500,
            "demand_supply_gap": 1500,
            "population": 579000,
            "housing_stock": 260400,
            "net_migration": 2600,
            "bungalow_demand": 1050,
            "bungalow_supply": 500,
            "population_makeup": {
                "age_distribution": {
                    "0-14": 84000,
                    "15-24": 66000,
                    "25-44": 143000,
                    "45-64": 146000,
                    "65+": 137000
                },
                "household_types": {
                    "single_person": 76000,
                    "couple_no_children": 60000,
                    "couple_with_children": 91000,
                    "single_parent": 41000,
                    "other": 26000
                },
                "tenure": {
                    "owner_occupied": 62,
                    "private_rented": 22,
                    "social_rented": 16
                },
                "ethnicity": {
                    "White": 90.2,
                    "Asian": 5.4,
                    "Black": 2.4,
                    "Mixed": 1,
                    "Other": 1
                }
            }
        }
    ]
}

# ==========================
# Extract Forecast Data
# ==========================
def extract_forecast():
    return FORECAST_JSON["forecast"]

# ==========================
# Excel File Generation (Updated for Bungalow Analysis)
# ==========================
def create_excel(forecast):
    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast Summary"

    summary_data = {
        "Year": [],
        "Predicted Demand": [],
        "Predicted Supply": [],
        "Bungalow Demand": [],
        "Bungalow Supply": [],
        "Demand-Supply Gap": [],
        "Population": [],
        "Housing Stock": [],
        "Net Migration": []
    }

    for year_data in forecast:
        bd = year_data.get("bungalow_demand", 0)
        bs = year_data.get("bungalow_supply", 0)
        summary_data["Year"].append(year_data["year"])
        summary_data["Predicted Demand"].append(year_data["predicted_demand"])
        summary_data["Predicted Supply"].append(year_data["predicted_supply"])
        summary_data["Bungalow Demand"].append(bd)
        summary_data["Bungalow Supply"].append(bs)
        summary_data["Demand-Supply Gap"].append(bd - bs)
        summary_data["Population"].append(year_data["population"])
        summary_data["Housing Stock"].append(year_data["housing_stock"])
        summary_data["Net Migration"].append(year_data["net_migration"])

    df = pd.DataFrame(summary_data)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Line Chart: Bungalow Demand vs Supply
    chart = LineChart()
    chart.title = "Bungalow Demand vs Supply"
    chart.y_axis.title = "Units"
    chart.x_axis.title = "Year"

    data = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=len(df) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I2")

    # Bar Chart: Demand-Supply Gap
    gap_chart = BarChart()
    gap_chart.title = "Bungalow Demand-Supply Gap"
    gap_chart.y_axis.title = "Gap (Units)"
    gap_chart.x_axis.title = "Year"

    gap_data = Reference(ws, min_col=6, max_col=6, min_row=1, max_row=len(df) + 1)
    gap_chart.add_data(gap_data, titles_from_data=True)
    gap_chart.set_categories(cats)
    ws.add_chart(gap_chart, "I20")

    # Age Group Distribution Chart (2025)
    age_groups = list(forecast[0]["population_makeup"]["age_distribution"].keys())
    age_data = [forecast[0]["population_makeup"]["age_distribution"][age] for age in age_groups]

    for idx, age_group in enumerate(age_groups, start=2):
        ws[f"I{idx}"] = age_group
        ws[f"J{idx}"] = age_data[idx - 2]

    age_chart = BarChart()
    age_chart.title = "Population Age Distribution (2025)"
    age_chart.y_axis.title = "Population"
    age_chart.x_axis.title = "Age Group"

    age_data_ref = Reference(ws, min_col=10, max_col=10, min_row=2, max_row=1 + len(age_groups))
    age_chart.add_data(age_data_ref, titles_from_data=True)
    age_chart.set_categories(Reference(ws, min_col=9, min_row=2, max_row=1 + len(age_groups)))
    ws.add_chart(age_chart, "L2")

    # Save file
    wb.save("bungalow_housing_forecast_analysis2.xlsx")
    print("✅ Excel file 'bungalow_housing_forecast_analysis2.xlsx' created.")
    return df

# ==========================
# Main Execution
# ==========================
if __name__ == "__main__":

    # Create Excel with tables and graphs
    forecast_data = extract_forecast()
    create_excel(forecast_data)